import argparse
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil import parser as dateparser

parser = argparse.ArgumentParser(description="Send vehicle data to server and write to Excel.")
parser.add_argument("-k", "--keys", nargs="*", default=[], help="Fields to include")
parser.add_argument("-c", "--colored", action="store_true", default=True, help="Enable colored rows based on HU")
args = parser.parse_args()

with open("vehicles.csv", "rb") as file:
    response = requests.post("http://127.0.0.1:5000/upload-csv", files={"file": file})

if response.status_code != 200:
    print("Server error:", response.text)
    exit()

data = response.json()
if isinstance(data, dict) and "message" in data:
    print("Server message:", data["message"])
    exit()

filtered_data = [item for item in data if isinstance(item, dict)]
sorted_data = sorted(filtered_data, key=lambda x: x.get("gruppe", ""))

wb = Workbook()
ws = wb.active
ws.title = "Vehicles"

headers = ["rnr"] + args.keys
ws.append(headers)

column_widths = {i: len(h) + 2 for i, h in enumerate(headers, start=1)}

COLOR_GREEN = "007500"
COLOR_ORANGE = "FFA500"
COLOR_RED = "b30000"

for item in sorted_data:
    row_values = [item.get("rnr", "")]

    for key in args.keys:
        if key == "labelIds":
            labels = item.get("resolved_labels", [])
            label_names = [lbl.get("label", "") for lbl in labels]
            joined_labels = ", ".join(label_names) if label_names else ""
            row_values.append(joined_labels)
        else:
            value = item.get(key, "")
            row_values.append(str(value) if value else "")

    ws.append(row_values)
    row_index = ws.max_row

    for col_idx, value in enumerate(row_values, start=1):
        val_len = len(str(value))
        if val_len > column_widths[col_idx]:
            column_widths[col_idx] = val_len

    if "labelIds" in args.keys:
        label_col_idx = headers.index("labelIds") + 1
        labels = item.get("resolved_labels", [])
        if labels:
            first_color = labels[0].get("color", "").replace("#", "")
            if first_color:
                cell = ws.cell(row=row_index, column=label_col_idx)
                cell.font = Font(color=first_color)

    if args.colored:
        hu_date = item.get("hu")
        if hu_date:
            try:
                hu = dateparser.parse(hu_date)
                today = datetime.today()
                months_diff = (today.year - hu.year) * 12 + (today.month - hu.month)

                if months_diff <= 3:
                    fill = PatternFill(start_color=COLOR_GREEN, fill_type="solid")
                elif months_diff <= 12:
                    fill = PatternFill(start_color=COLOR_ORANGE, fill_type="solid")
                else:
                    fill = PatternFill(start_color=COLOR_RED, fill_type="solid")

                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_index, column=col_idx)
                    cell.fill = fill
            except:
                pass

for i, width in column_widths.items():
    col_letter = get_column_letter(i)
    ws.column_dimensions[col_letter].width = width

filename = f"vehicles_{datetime.today().date().isoformat()}.xlsx"
wb.save(filename)
print(f"Create Excel File: {filename}")
